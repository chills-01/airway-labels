# file that handles the labelling section of the program
from anytree import PreOrderIter
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import os


def has_false_bifurcations(tree):
    false_nodes = []
    for node in PreOrderIter(tree):
        if len(node.children) > 2:
            false_nodes.append(node)
    
    if len(false_nodes) != 0:
        return false_nodes
    else:
        return None


def label_tree(tree, endpoints):
    """
    Takes the tree object and endpoint coordinates, and then returns the labels
    T, RMB, LMB, R1, R2, R3, R4 with corresponding number.
    Done in order of depth level - may need to be chaged depending on diagram Melissa uses
    This diagram as a reference: https://www.spiedigitallibrary.org/ContentImages/Proceedings/11313/1131312/FigureImages/00029_PSISDG11313_1131312_page_4_1.jpg
    """

    names = ['Trachea', 'RMB', 'LMB', 'R1', 'R2', 'R3', 'R4']
    #initialise dictionary:
    labels = {name: None for name in names}

    # trachea
    labels[names[0]] = tree.name
    names.pop(0)

    # RMB and LMB
    children = tree.children
    children_avg_coords = []
    for child in children:
        name = child.name
        endpoint_coords = endpoints[name]
        avg_coords = [(g + h) / 2 for g, h in zip(endpoint_coords[0], endpoint_coords[1])]
        children_avg_coords.append((child, avg_coords))

    # left side has bigger y coordinate
    b1, b2 = children_avg_coords[0][1], children_avg_coords[1][1] 
    y1, y2 = b1[1], b2[1]

    if y1 > y2:
        lmb = children_avg_coords[0][0]  
        rmb = children_avg_coords[1][0]
    else:
        lmb = children_avg_coords[1][0]  
        rmb = children_avg_coords[0][0]

    labels[names[0]] = rmb.name
    names.pop(0)
    labels[names[0]] = lmb.name
    names.pop(0)

    # right upper and lower
    children = rmb.children
    children_avg_coords = []
    for child in children:
        name = child.name
        endpoint_coords = endpoints[name]
        avg_coords = [(g + h) / 2 for g, h in zip(endpoint_coords[0], endpoint_coords[1])]
        children_avg_coords.append((child, avg_coords))

    # determine the upper/lower bifurcation on the right side
    b1, b2 = children_avg_coords[0][1], children_avg_coords[1][1] 
    z1, z2 = b1[2], b2[2]

    if z1 < z2:
        upper = children_avg_coords[0][0]  
        lower = children_avg_coords[1][0]
    else:
        upper = children_avg_coords[1][0]  
        lower = children_avg_coords[0][0]
    


    children = upper.children
    children_avg_coords = []
    for child in children:
        name = child.name
        endpoint_coords = endpoints[name]
        avg_coords = [(g + h) / 2 for g, h in zip(endpoint_coords[0], endpoint_coords[1])]
        children_avg_coords.append((child, avg_coords))
     
    # r1 has smaller z coordinate
    b1, b2 = children_avg_coords[0][1], children_avg_coords[1][1] 
    z1, z2 = b1[2], b2[2]

    if z1 < z2:
        r1 = children_avg_coords[0][0]  
        r23 = children_avg_coords[1][0]
    else:
        r1 = children_avg_coords[1][0]  
        r23 = children_avg_coords[0][0]

    labels[names[0]] = r1.name
    names.pop(0)

    # r2 and r3 branches off r23 (r2 has lower z)
    children = upper.children
    children_avg_coords = []
    for child in children:
        name = child.name
        endpoint_coords = endpoints[name]
        avg_coords = [(g + h) / 2 for g, h in zip(endpoint_coords[0], endpoint_coords[1])]
        children_avg_coords.append((child, avg_coords))
     
    # r1 has smaller z coordinate

    children = r23.children
    children_avg_coords = []
    for child in children:
        name = child.name
        endpoint_coords = endpoints[name]
        avg_coords = [(g + h) / 2 for g, h in zip(endpoint_coords[0], endpoint_coords[1])]
        children_avg_coords.append((child, avg_coords))
    
    b1, b2 = children_avg_coords[0][1], children_avg_coords[1][1] 
    z1, z2 = b1[2], b2[2]

    if z1 < z2:
        r2 = children_avg_coords[0][0]  
        r3 = children_avg_coords[1][0]
    else:
        r2 = children_avg_coords[1][0]  
        r3 = children_avg_coords[0][0]

    labels[names[0]] = r2.name
    names.pop(0)

    labels[names[0]] = r3.name
    names.pop(0)

    # r4: along lower, then r45 then r4 lies below r5

    return labels
    

def write_to_excel(labelled_tree, airway_name):

    excel_name = 'output.xlsx'
    # put all the headings in
    wb = Workbook()
    ws = wb.active
    row = 2
    col = 2
    
    # heading
    ws.cell(row, col, value = airway_name)
    ws.cell(row,col).font = Font(italic=True, bold=True)

    
    orig_row = row + 1
    for key, value in labelled_tree.items():

        if value is None:
            value = 'Undetermined'
        row = orig_row
        ws.cell(row, col, value = key)
        ws.cell(row, col).font = Font(bold=True)

        row += 1
        ws.cell(row, col, value = value)

        row = orig_row
        col += 1

    if excel_name in os.listdir('.'):
        try:
            name = excel_name.split('.')[0]
            number = int(name[-1])
        except ValueError:
            number = 2
        
        excel_name = 'output_' + str(number) + '.xlsx'


    wb.save(excel_name)